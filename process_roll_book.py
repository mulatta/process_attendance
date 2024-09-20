import pandas as pd
import os
import argparse
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def process_attendance(rb, attendance_df, round_num):
    outlier = []
    attendance_column = f'{round_num}차 출결'

    for _, row in attendance_df.iterrows():
        id = row[attendance_df.columns[2]]
        dept = row[attendance_df.columns[3]].replace('수도권 ', '')
        email = row[attendance_df.columns[1]]
        
        found = False
        for sheet_name, sheet_df in rb.items():
            mask = (sheet_df['이름+전화번호 뒤 4자리'] == id)
            if mask.any():
                rb[sheet_name].loc[mask, attendance_column] = 1
                found = True
                break
        
        if not found:
            outlier.append({
                '이메일 주소': email,
                '이름+전화번호 뒤 4자리': id,
                '소속분과': dept
            })
            print(f"{attendance_column}: ID {id}에 해당하는 사용자를 찾을 수 없습니다.")

    # NaN 값을 0으로 채우기
    for k in rb.keys():
        rb[k][attendance_column] = rb[k][attendance_column].fillna(0)

    return outlier

def save_results(rb, filename):
    os.makedirs(os.path.dirname(filename), exist_ok=True)
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        for sheet_name, df in rb.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    print(f"결과가 {filename}에 저장되었습니다.")

def save_outliers(outliers, filename):
    os.makedirs(os.path.dirname(filename), exist_ok=True)
    df = pd.DataFrame(outliers)
    df.to_excel(filename, index=False)
    print(f"Outlier가 {filename}에 저장되었습니다.")

def apply_excel_formulas(filename):
    wb = load_workbook(filename)
    
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        max_row = ws.max_row
        max_col = ws.max_column
        
        # 출결 열 찾기
        attendance_cols = []
        for col in range(1, max_col + 1):
            cell_value = ws.cell(row=1, column=col).value
            if isinstance(cell_value, str) and '차 출결' in cell_value:
                attendance_cols.append(col)
        
        if not attendance_cols:
            print(f"'{sheet}' 시트에서 출결 열을 찾을 수 없습니다.")
            continue
        
        # '출결 횟수' 열 추가 및 함수식 적용
        attendance_count_col = max_col + 1
        ws.cell(row=1, column=attendance_count_col, value='출결 횟수')
        for row in range(2, max_row + 1):
            formula = f'=SUM({",".join([f"{get_column_letter(col)}{row}" for col in attendance_cols])})'
            ws.cell(row=row, column=attendance_count_col, value=formula)
        
        # '최종 출결' 열 추가 및 함수식 적용
        final_attendance_col = max_col + 2
        ws.cell(row=1, column=final_attendance_col, value='최종 출결')
        attendance_count_col_letter = get_column_letter(attendance_count_col)
        total_sessions = len(attendance_cols)
        
        for row in range(2, max_row + 1):
            conditions = []
            conditions.append(f'{attendance_count_col_letter}{row}>={total_sessions-1}')  # 전체 회차의 1회 미만 결석
            
            # 1, 2차 연속 결석이나 마지막 두 차 연속 결석 시 5회 출석이면 출석으로 인정
            if total_sessions >= 2:
                first_two = f'AND({attendance_count_col_letter}{row}={total_sessions-2},{get_column_letter(attendance_cols[0])}{row}=0,{get_column_letter(attendance_cols[1])}{row}=0)'
                last_two = f'AND({attendance_count_col_letter}{row}={total_sessions-2},{get_column_letter(attendance_cols[-2])}{row}=0,{get_column_letter(attendance_cols[-1])}{row}=0)'
                conditions.append(first_two)
                conditions.append(last_two)
            
            formula = f'=IF(OR({",".join(conditions)}),"출석","결석")'
            ws.cell(row=row, column=final_attendance_col, value=formula)
        
        # '수료 여부' 열 업데이트
        previous_completion_col = None
        for col in range(1, max_col + 1):
            if ws.cell(row=1, column=col).value == '수료 여부':
                previous_completion_col = col
                ws.cell(row=1, column=col, value='최종 발표 이전 수료 여부')
                break
        
        if previous_completion_col is None:
            print(f"'{sheet}' 시트에서 '수료 여부' 열을 찾을 수 없습니다.")
            continue
        
        final_completion_col = max_col + 3
        ws.cell(row=1, column=final_completion_col, value='수료 여부')
        previous_completion_col_letter = get_column_letter(previous_completion_col)
        final_attendance_col_letter = get_column_letter(final_attendance_col)
        
        for row in range(2, max_row + 1):
            formula = f'=IF({previous_completion_col_letter}{row}="수료", "수료", ' \
                      f'IF(AND({previous_completion_col_letter}{row}="수료 예정", ' \
                      f'{final_attendance_col_letter}{row}="출석"), "수료", "수료 불가"))'
            ws.cell(row=row, column=final_completion_col, value=formula)
    
    wb.save(filename)
    print(f"Excel 함수식이 {filename}에 적용되었습니다.")

def main():
    parser = argparse.ArgumentParser(description='출석 처리 스크립트')
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument('--attendance', type=int, help='처리할 출석 회차')
    group.add_argument('--results', action='store_true', help='최종 결과 생성')
    args = parser.parse_args()

    if args.attendance:
        # 사용자 정보 파일 읽기
        rb = pd.read_excel('user_info/users.xlsx',
                            sheet_name=['1분과', '2분과', '3분과', '4분과', '충청분과', '영남분과'])

        # 이전 회차의 처리 결과 확인
        prev_result_file = f'./processed_attendance/processed_attendance_{args.attendance-1}.xlsx'
        if args.attendance > 1 and os.path.exists(prev_result_file):
            print(f"{args.attendance-1}차 처리 결과를 불러옵니다.")
            rb = pd.read_excel(prev_result_file, sheet_name=None)

        # 출결 파일 처리
        attendance_file = f'./attendance_forms/{args.attendance}차 출결.xlsx'
        if os.path.exists(attendance_file):
            print(f"{args.attendance}차 출결 파일 처리 중...")
            attendance_df = pd.read_excel(attendance_file)
            outliers = process_attendance(rb, attendance_df, args.attendance)
            
            # 결과 저장
            save_results(rb, f'./processed_attendance/processed_attendance_{args.attendance}.xlsx')
            save_outliers(outliers, f'./outliers/outliers_{args.attendance}.xlsx')
        else:
            print(f"{args.attendance}차 출결 파일이 존재하지 않습니다.")

    elif args.results:
        # 모든 처리된 파일 읽기
        processed_files = sorted([f for f in os.listdir('./processed_attendance') if f.startswith('processed_attendance_')])
        if not processed_files:
            print("처리된 출석 파일이 없습니다.")
            return

        latest_file = os.path.join('./processed_attendance', processed_files[-1])
        rb = pd.read_excel(latest_file, sheet_name=None)

        # 최종 결과 파일 생성 및 Excel 함수식 적용
        final_result_file = './results/final_attendance_result.xlsx'
        save_results(rb, final_result_file)
        apply_excel_formulas(final_result_file)

    print("처리가 완료되었습니다.")

if __name__ == "__main__":
    main()